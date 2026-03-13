import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
FILE_PATH = "data/products.xlsx"
LOW_STOCK = 2
class ProductService:
    def __init__(self):
        self.file = FILE_PATH
        if not os.path.exists(self.file):
            self.create_file()
    def create_file(self):
        os.makedirs("data", exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.append(["Tên sản phẩm", "Size", "Màu sắc", "Số lượng", "Giá"])
        wb.save(self.file)
    def get_products(self):
        df = pd.read_excel(self.file)
        products = []
        for _, row in df.iterrows():
            product = {
                "name": row["Tên sản phẩm"],
                "size": str(row["Size"]),
                "color": row["Màu sắc"],
                "quantity": int(row["Số lượng"]),
                "price": int(row["Giá"])
            }
            products.append(product)
        return products, []
    def add_product(self, product):
        df = pd.read_excel(self.file)
        new_row = {
            "Tên sản phẩm": product["name"],
            "Size": product["size"],
            "Màu sắc": product["color"],
            "Số lượng": product["quantity"],
            "Giá": product["price"]
        }
        df.loc[len(df)] = new_row
        df.to_excel(self.file, index=False)
    def update_product(self, row, product):
        df = pd.read_excel(self.file)
        df.loc[row] = [
            product["name"],
            product["size"],
            product["color"],
            product["quantity"],
            product["price"]
        ]
        df.to_excel(self.file, index=False)
    def delete_product(self, index):
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        ws.delete_rows(index + 2)
        wb.save(FILE_PATH)
